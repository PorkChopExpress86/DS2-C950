import csv

import openpyxl

"""This will convert the Excel files in the Excel folder into CSV and clean the data. This only needed to be run once and was not needed
to determine the route for the truck"""


def write_to_csv(output_path, list_to_write):
    with open(output_path, "w") as f:
        writer = csv.writer(f, lineterminator="\n")
        writer.writerows(list_to_write)


def packages_to_csv():
    wb = openpyxl.load_workbook(r"ExcelFiles/WGUPS Package File.xlsx")
    ws = wb.active
    package_addresses = []

    for row in ws.iter_rows(min_row=9, max_row=48, min_col=2, max_col=8):
        package_address = []
        for cell in row:
            cell_str = (
                str(cell.value)
                .replace("\r", "")
                .replace("\n", "")
                .replace("South ", "S")
                .replace("North ", "N")
                .replace("East ", "E")
                .replace("West ", "W")
            )
            package_address.append(cell_str)
        package_addresses.append(package_address)

    write_to_csv("CSVFiles/packages.csv", package_addresses)


def addresses_to_csv():
    wb = openpyxl.load_workbook(r"ExcelFiles/WGUPS Distance Table.xlsx")
    ws = wb.active

    addresses = []

    for row in ws.iter_rows(min_row=9, max_row=35, min_col=1, max_col=1):
        address = []
        for cell in row:
            cell_list = (
                str(cell.value)
                .replace("\r", "")
                .replace(",", "")
                .replace("South ", "S")
                .replace("North ", "N")
                .replace("East ", "E")
                .replace("West ", "W")
            ).split("\n")
            for index, item in enumerate(cell_list):
                cell_list[index] = item.strip()

            addresses.append(cell_list)

    write_to_csv("CSVFiles/addresses.csv", addresses)


def distance_table_to_csv():
    wb = openpyxl.load_workbook(r"ExcelFiles/WGUPS Distance Table.xlsx")
    ws = wb.active

    distances = []

    for row in ws.iter_rows(min_row=9, max_row=35, min_col=3, max_col=29):
        distance_row = []
        for cell in row:
            distance = str(cell.value)
            distance_row.append(distance)
        distances.append(distance_row)

    write_to_csv("CSVFiles/distance_table.csv", distances)


packages_to_csv()
addresses_to_csv()
distance_table_to_csv()
